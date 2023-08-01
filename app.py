import os
import streamlit as st
import openpyxl

# ... Rest of the code ...

def classify_grade(value):
    if value <= 0:
        return "no stenosis"
    elif value > 0 and value <= 50:
        return "Grade 1"
    elif value >= 51 and value <= 70:
        return "Grade 2"
    elif value >= 71 and value <= 99:
        return "Grade 3"
    else:
        return "Out of Range"

def main():
    # Load the Excel file using relative path
    wb = openpyxl.load_workbook('evan propst sizes.xlsx')

    # Create a dropdown menu for the sheet name
    sheet_name = st.selectbox('ET Tube Manufacturer', wb.sheetnames)

    # Get the selected sheet
    sheet = wb[sheet_name]

    # Get the x-axis and y-axis labels
    x_axis_labels = [cell.value for cell in sheet[2]][1:]

    # Get the Y-axis labels by iterating through the first column (excluding the header)
    y_axis_labels = [cell.value for col in sheet.iter_cols(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1) for cell in col][1:]

    # Create a dropdown menu for the y-axis label
    y_axis_label = st.selectbox('Age', y_axis_labels)

    # Create a dropdown menu for the x-axis label
    x_axis_label = st.selectbox('Actual measured tube size', x_axis_labels)

    # Get the row and column indices of the selected labels
    x_axis_index = x_axis_labels.index(x_axis_label) + 2
    y_axis_index = y_axis_labels.index(y_axis_label) + 3

    # Get the value of the cell
    value = sheet.cell(row=y_axis_index, column=x_axis_index).value

    # Classify the value into grades
    grade = classify_grade(value)

    # Display the value
    st.write(f"Subglottic stenosis %: {value}")
    st.write(f"Cotton Myer Grade: {grade}")
    st.write("Source: Propst, E.J. and Wolter, N.E. (2023), Myer-Cotton Grade of Subglottic Stenosis Depends on Style of Endotracheal Tube Used. The Laryngoscope. https://doi.org/10.1002/lary.30577")
    st.write("Adapted by: Darth Vader a.k.a Eishaan Bhargava https://twitter.com/darthekb")

image = Image.open('vader.jpg')

st.image(image, caption='Adapted by: Darth Vader a.k.a Eishaan Bhargava https://twitter.com/darthekb')

if __name__ == '__main__':
    # Use the os module to dynamically set the port
    port = int(os.environ.get('PORT', 8501))
    main()
